#!/usr/bin/env python3
# Rebuild the ATC Site of Care dashboard from REAL data.
# 1) Run the query in "Patient Data query (Snowflake).sql", export the grid as CSV.
# 2) Save that file next to this script as  patient_data.csv  (keep the 17 column headers).
# 3) Run:  python3 build_workbook.py
# It writes "ATC Site of Care - DASHBOARD.xlsx" in this folder, with the account lists,
# regions, charts and KPIs all rebuilt from your data. Needs: pip install openpyxl
#!/usr/bin/env python3
"""
Build: ATC Site of Care - DASHBOARD.xlsx
A top-firm-style, slide-led, interactive Excel dashboard. One 'Patient Data' sheet
drives every slide via live formulas (COUNTIFS/SUMIFS/AVERAGEIFS). Seeded with a
validated sample so it renders truthfully out of the box; user replaces the data.
"""
import csv, statistics
from collections import Counter, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, DoughnutChart, BarChart, LineChart, ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule

import os
HERE=os.path.dirname(os.path.abspath(__file__))
OUT=os.path.join(HERE,"ATC Site of Care - DASHBOARD.xlsx")

# ---------------- palette ----------------
NAVY="1F3864"; INK="212B36"; GREEN="2E7D32"; GREEN_D="1B5E20"; SLATE="8497B0"; SLATE_D="5B6B85"
AMBER="C9A227"; TEAL="0E7C7B"; PAPER="F5F7FA"; CARD="FFFFFF"; LINE="D9DEE6"; MUTE="6B7280"
WHITE="FFFFFF"; INPUT="FFF3CC"; GREENBG="E8F1E9"; NAVYBG="E7ECF5"
FONT="Arial"
thin=Side(style="thin",color=LINE)
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
CNT='#,##0'; PCT='0.0%'; PCT0='0%'; AV='0.0'; INT='0'; DAY='0" d"'

def Fnt(sz=10,b=False,c=INK,i=False): return Font(name=FONT,size=sz,bold=b,color=c,italic=i)
def Fill(h): return PatternFill("solid",fgColor=h)
CEN=Alignment(horizontal="center",vertical="center",wrap_text=True)
LEF=Alignment(horizontal="left",vertical="center",wrap_text=True)
RGT=Alignment(horizontal="right",vertical="center")
TOP=Alignment(horizontal="left",vertical="top",wrap_text=True)

# ---------------- load sample ----------------
rows=list(csv.DictReader(open(os.path.join(HERE,"patient_data.csv"))))
for r in rows:
    for k in ("is_atc","first_year","started_atc","days_dx_to_tx","treatment_claims","yervoy","opdualag"):
        r[k]=int(r[k])
N=len(rows)

REGIONS=["West","Central","Great Lakes","Ohio Valley","Southeast","Northeast"]
BUCKETS=["ATC","Non-ATC: Hospital","Non-ATC: Community network","Non-ATC: Other"]
CLASSES=["ATC: NPI confirmed","ATC: roster gap corrected","ATC: name fallback","Non-ATC",
         "Non-ATC: System sweep","Non-ATC: Community Network","Non-ATC: Unknown","Needs Review"]

# seed account lists from sample
atc_parent_ct=Counter(r["account_parent"] for r in rows if r["is_atc"]==1)
atc_parents=[p for p,_ in atc_parent_ct.most_common()]
# "genuine targets" for slides 7 & 8: drop community networks and the big multi-site
# systems that already sit on the ATC roster with one authorized site (matches deck B7B/B8B).
EXCLUDE=["US ONCOLOGY","ONE ONCOLOGY","AMERICAN ONCOLOGY","KAISER","PROVIDENCE","MAYO",
         "INTERMOUNTAIN","AVERA","NORTHWELL","ADVENTHEALTH","ADVOCATE","ST LUKE","BAYLOR"]
def genuine(name): return not any(x in name.upper() for x in EXCLUDE)
non_by_region=defaultdict(Counter)
for r in rows:
    if r["is_atc"]==0 and genuine(r["account_parent"]): non_by_region[r["region"]][r["account_parent"]]+=1
non_by_state=defaultdict(Counter)
for r in rows:
    if r["is_atc"]==0 and r["region"]!="Unmapped" and genuine(r["account_parent"]): non_by_state[r["state"]][r["account_parent"]]+=1
state_untapped=Counter()
for r in rows:
    if r["is_atc"]==0 and r["region"]!="Unmapped": state_untapped[r["state"]]+=1
top_states=[s for s,_ in state_untapped.most_common(8)]

wb=openpyxl.Workbook()

# ================= column handles on Patient Data =================
PD="'Patient Data'"
A=lambda col:f"{PD}!${col}:${col}"
C_ID,C_BUCKET,C_ATC,C_CLASS,C_MATCH,C_PARENT,C_COMM,C_STATE,C_REGION,C_YEAR,C_STARTED,C_FIRST,C_LAST,C_DAYS,C_CLAIMS,C_YERVOY,C_OPD = \
 [A(c) for c in "ABCDEFGHIJKLMNOPQ"]
TOT=f"COUNT({C_ATC})"; ATCN=f"COUNTIF({C_ATC},1)"; NONATCN=f"COUNTIF({C_ATC},0)"

# ================= generic helpers =================
def ws_new(name,tab=NAVY,hide_grid=True):
    ws=wb.create_sheet(name)
    if hide_grid: ws.sheet_view.showGridLines=False
    ws.sheet_properties.tabColor=tab
    return ws
def band(ws,r,c1,c2,hexc):
    for c in range(c1,c2+1): ws.cell(row=r,column=c).fill=Fill(hexc)
def hero(ws,r,text,sub,c2=12):
    ws.merge_cells(start_row=r,start_column=1,end_row=r+2,end_column=c2)
    band(ws,r,1,c2,NAVY); band(ws,r+1,1,c2,NAVY); band(ws,r+2,1,c2,NAVY)
    cell=ws.cell(row=r,column=1,value=text); cell.font=Fnt(20,True,WHITE); cell.alignment=Alignment(horizontal="left",vertical="center")
    s=ws.cell(row=r+2,column=1); # sub on its own merged strip below
    return r+3
def section(ws,r,text,c2=12,color=NAVY):
    ws.cell(row=r,column=1,value=text).font=Fnt(13,True,color)
    ws.cell(row=r+1,column=1,value=None)
    band(ws,r+1,1,c2,color)   # thin underline strip
    ws.row_dimensions[r+1].height=3
    return r+2
def kpi_card(ws,r,c,value_formula,label,accent,numfmt=CNT,note=None,w=3):
    """3-row card: accent strip (row r), big number (r+1), label (r+2). Kept deliberately spare."""
    ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+w-1)          # accent strip
    band(ws,r,c,c+w-1,accent); ws.row_dimensions[r].height=6
    for rr in (r+1,r+2):
        for cc in range(c,c+w): ws.cell(row=rr,column=cc).fill=Fill(CARD)
    for rr in range(r,r+3): band_cells_border(ws,rr,c,c+w-1)
    ws.merge_cells(start_row=r+1,start_column=c,end_row=r+1,end_column=c+w-1)
    v=ws.cell(row=r+1,column=c,value=value_formula); v.font=Fnt(26,True,accent); v.alignment=CEN; v.number_format=numfmt
    ws.merge_cells(start_row=r+2,start_column=c,end_row=r+2,end_column=c+w-1)
    l=ws.cell(row=r+2,column=c,value=label); l.font=Fnt(10,True,MUTE); l.alignment=CEN
    ws.row_dimensions[r+1].height=34
def band_cells_border(ws,r,c1,c2):
    side=Side(style="thin",color=LINE)
    for c in range(c1,c2+1):
        cur=ws.cell(row=r,column=c)
        cur.border=Border(left=side,right=side,top=side,bottom=side)
def th(ws,r,headers,c0=1,fill=NAVY):
    for i,h in enumerate(headers):
        x=ws.cell(row=r,column=c0+i,value=h); x.fill=Fill(fill); x.font=Fnt(10,True,WHITE); x.alignment=CEN; x.border=BORDER
    ws.row_dimensions[r].height=26
def td(ws,r,c,val,fmt=None,b=False,al=RGT,fill=None,color=INK,border=True):
    x=ws.cell(row=r,column=c,value=val); x.font=Fnt(10,b,color); x.alignment=al
    if border: x.border=BORDER
    if fmt: x.number_format=fmt
    if fill: x.fill=Fill(fill)
    return x
def widths(ws,m):
    for k,v in m.items(): ws.column_dimensions[k].width=v
def note(ws,r,text,c2=12,color=MUTE,size=9):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=c2)
    x=ws.cell(row=r,column=1,value=text); x.font=Fnt(size,False,color,True); x.alignment=TOP
    ws.row_dimensions[r].height=max(14,12*(len(text)//120+1))
    return r+1

def style_series(chart, colors):
    for i,s in enumerate(chart.series):
        col=colors[i%len(colors)]
        s.graphicalProperties.solidFill=col
        s.graphicalProperties.line.solidFill=col

# ================================================================ helper: paper background
def paper(ws,rows_=80,cols_=14):
    for r in range(1,rows_):
        for c in range(1,cols_):
            if ws.cell(row=r,column=c).fill.patternType is None:
                ws.cell(row=r,column=c).fill=Fill(PAPER)

print("scaffolding sheets...")
# create sheets up front for hyperlinks/order
cover=wb.active; cover.title="Cover"; cover.sheet_view.showGridLines=False; cover.sheet_properties.tabColor=NAVY
dump=ws_new("ATC Patient Counts",GREEN_D)
pdata=ws_new("Patient Data",GREEN_D)
loadsql=ws_new("Load Data (SQL)",GREEN_D)
rmap=ws_new("Region Map",SLATE_D)
s3=ws_new("S3 · Market Structure",NAVY)
s4=ws_new("S4 · Patient Journey",NAVY)
s5=ws_new("S5 · Regional Penetration",NAVY)
s6=ws_new("S6 · State Scatter",NAVY)
s7=ws_new("S7 · Non-ATC by Region",AMBER)
s8=ws_new("S8 · Non-ATC by State",AMBER)
s9=ws_new("S9 · Appendix",TEAL)
meth=ws_new("Methodology",SLATE_D)

# ================================================================ PATIENT DATA
widths(pdata,{"A":12,"B":22,"C":8,"D":24,"E":15,"F":32,"G":24,"H":8,"I":13,"J":10,"K":10,"L":11,"M":11,"N":13,"O":13,"P":9,"Q":10})
band(pdata,1,1,17,GREEN_D)
pdata.merge_cells("A1:Q1")
pdata["A1"]="PATIENT DATA  —  one row per patient  ·  this is the engine; every slide reads from here"
pdata["A1"].font=Fnt(12,True,WHITE); pdata["A1"].alignment=Alignment(horizontal="left",vertical="center")
pdata.row_dimensions[1].height=24
pdata.merge_cells("A2:Q2")
pdata["A2"]=("This is sample data, built to match the deck so you can see the workbook working. "
             "To use the real numbers: clear everything below the header row and paste your Snowflake export starting in A4 — keep the column names exactly as they are.")
pdata["A2"].font=Fnt(9,False,"9C6B00",True); pdata["A2"].fill=Fill(INPUT); pdata["A2"].alignment=TOP
pdata.row_dimensions[2].height=28
HEAD=["patient_id","site_bucket","is_atc","class_hybrid","match_basis","account_parent","community_network",
      "state","region","first_year","started_atc","first_site","last_site","days_dx_to_tx","treatment_claims","yervoy","opdualag"]
th(pdata,3,HEAD)
pdata.freeze_panes="A4"
cols=["patient_id","site_bucket","is_atc","class_hybrid","match_basis","account_parent","community_network",
      "state","region","first_year","started_atc","first_site","last_site","days_dx_to_tx","treatment_claims","yervoy","opdualag"]
numeric={2,9,10,13,14,15,16}  # 0-based indexes that are numeric
for ri,r in enumerate(rows):
    excel_r=4+ri
    for ci,key in enumerate(cols):
        v=r[key]
        c=pdata.cell(row=excel_r,column=ci+1,value=v)
        c.font=Fnt(9); c.alignment=CEN if ci not in (5,6,3) else LEF
        if ci in (2,10): c.number_format=INT
# light banding via alt fill on first few won't scale; skip for perf. Add autofilter.
pdata.auto_filter.ref=f"A3:Q{3+N}"
print("patient data written:",N,"rows")

# ================================================================ REGION MAP
widths(rmap,{"A":10,"B":16})
band(rmap,1,1,2,SLATE_D); rmap.merge_cells("A1:B1")
rmap["A1"]="STATE → REGION"; rmap["A1"].font=Fnt(12,True,WHITE)
th(rmap,3,["State","Region"])
REGION_MAP={}
import importlib.util
# rebuild map same as generator
_m={}
for st in "CA WA OR NV AZ UT CO ID MT WY NM".split(): _m[st]="West"
for st in "TX OK KS NE SD ND AR".split(): _m[st]="Central"
for st in "IL MI WI MN IA MO".split(): _m[st]="Great Lakes"
for st in "OH IN KY TN WV".split(): _m[st]="Ohio Valley"
for st in "FL GA SC NC AL MS LA".split(): _m[st]="Southeast"
for st in "NY NJ PA MA CT RI NH ME VT VA MD DC DE".split(): _m[st]="Northeast"
rr=4
for st in sorted(_m):
    td(rmap,rr,1,st,al=CEN); td(rmap,rr,2,_m[st],al=LEF); rr+=1
rmap.freeze_panes="A4"
note(rmap,rr+1,"Reference lookup used to roll states up to regions (from the Snowflake model). VA/MD/DC/DE sit in Northeast; HI/AK left unmapped.",c2=2)

print("region map done")

# ================================================================ LOAD DATA (SQL)
ws=loadsql; widths(ws,{"A":112})
band(ws,1,1,1,GREEN_D); ws["A1"]="LOAD DATA — one query builds the Patient Data sheet"
ws["A1"].font=Fnt(13,True,WHITE); ws.row_dimensions[1].height=26
guide=[
 "",
 "Run this in Snowflake, export the grid, and paste it into the 'Patient Data' sheet starting at cell A4",
 "(keep the column headers in row 3). Every slide, KPI and chart then recalculates automatically.",
 "It rebuilds the 17 patient-level columns from the base tables in the master file — no manual steps.",
 "",
]
r=2
for g in guide:
    c=ws.cell(row=r,column=1,value=g); c.font=Fnt(10,False,MUTE,True); c.alignment=LEF; r+=1
SQL=r'''WITH claim_roll AS (
    SELECT D_PATIENT_ID,
           MIN(YEAR(DATE_OF_SERVICE))                                       AS FIRST_YEAR,
           COUNT(*)                                                         AS TREATMENT_CLAIMS,
           MAX(CASE WHEN DRUG = 'Yervoy'   THEN 1 ELSE 0 END)               AS YERVOY,
           MAX(CASE WHEN DRUG = 'Opdualag' THEN 1 ELSE 0 END)               AS OPDUALAG,
           DATEDIFF('day', MIN(FIRST_DX_DATE), MIN(DATE_OF_SERVICE))        AS DAYS_DX_TO_TX
    FROM COMPILE_DEV.PUBLIC.ATC_TREATMENT_CLAIMS
    GROUP BY 1
),
first_last AS (
    SELECT D_PATIENT_ID,
           MAX(CASE WHEN rn_first = 1 THEN IS_ATC_HCO END) AS FIRST_ATC,
           MAX(CASE WHEN rn_last  = 1 THEN IS_ATC_HCO END) AS LAST_ATC
    FROM (
        SELECT D_PATIENT_ID, IS_ATC_HCO,
               ROW_NUMBER() OVER (PARTITION BY D_PATIENT_ID ORDER BY DATE_OF_SERVICE ASC,  D_PRIMARY_HCO_COMPILE_ID) AS rn_first,
               ROW_NUMBER() OVER (PARTITION BY D_PATIENT_ID ORDER BY DATE_OF_SERVICE DESC, D_PRIMARY_HCO_COMPILE_ID) AS rn_last
        FROM COMPILE_DEV.PUBLIC.ATC_TREATMENT_CLAIMS
    )
    GROUP BY 1
)
SELECT
    c.D_PATIENT_ID                                                          AS patient_id,
    CASE WHEN c.CLASS_FINAL = 'ATC'                                  THEN 'ATC'
         WHEN c.CLASS_FINAL = 'Non-ATC: Community Network'           THEN 'Non-ATC: Community network'
         WHEN c.CLASS_FINAL IN ('Non-ATC: Unknown','Needs Review')   THEN 'Non-ATC: Other'
         ELSE 'Non-ATC: Hospital' END                                      AS site_bucket,
    CASE WHEN c.CLASS_FINAL = 'ATC' THEN 1 ELSE 0 END                      AS is_atc,
    c.CLASS_HYBRID                                                          AS class_hybrid,
    CASE WHEN c.CLASS_HYBRID = 'ATC: NPI confirmed'        THEN 'NPI-confirmed'
         WHEN c.CLASS_HYBRID = 'ATC: roster gap corrected' THEN 'Roster-confirmed'
         WHEN c.CLASS_HYBRID = 'ATC: name fallback'        THEN 'Name-matched'
         ELSE '-' END                                                      AS match_basis,
    COALESCE(NULLIF(TRIM(c.HCO_PARENT_NAME), ''), 'Unknown / unmapped')     AS account_parent,
    COALESCE(c.HCO_COMMUNITY_NETWORK,
             CASE WHEN c.CLASS_FINAL = 'ATC' THEN '-' ELSE 'Independent / Other' END) AS community_network,
    c.PRIMARY_HCO_NPI_STATE                                                 AS state,
    COALESCE(r.REGION, 'Unmapped')                                         AS region,
    cr.FIRST_YEAR                                                          AS first_year,
    fl.FIRST_ATC                                                           AS started_atc,
    CASE WHEN fl.FIRST_ATC = 1 THEN 'ATC' ELSE 'Non-ATC' END              AS first_site,
    CASE WHEN fl.LAST_ATC  = 1 THEN 'ATC' ELSE 'Non-ATC' END              AS last_site,
    cr.DAYS_DX_TO_TX                                                       AS days_dx_to_tx,
    COALESCE(cr.TREATMENT_CLAIMS, 0)                                       AS treatment_claims,
    COALESCE(cr.YERVOY, 0)                                                 AS yervoy,
    COALESCE(cr.OPDUALAG, 0)                                               AS opdualag
FROM COMPILE_DEV.PUBLIC.ATC_CLASSIFIED_FINAL c
LEFT JOIN COMPILE_DEV.PUBLIC.STATE_REGION_MAP r ON c.PRIMARY_HCO_NPI_STATE = r.STATE
LEFT JOIN claim_roll  cr ON c.D_PATIENT_ID = cr.D_PATIENT_ID
LEFT JOIN first_last  fl ON c.D_PATIENT_ID = fl.D_PATIENT_ID;'''
for line in SQL.split("\n"):
    c=ws.cell(row=r,column=1,value=line if line else " ")
    c.font=Font(name="Consolas",size=9,color=INK); c.alignment=Alignment(horizontal="left",vertical="center")
    r+=1
ws.sheet_view.showGridLines=False
print("load-sql done")

# ================================================================ shared builders
from openpyxl.chart.series import DataPoint
def titleband(ws,text,sub,c2=12):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=c2); band(ws,1,1,c2,NAVY)
    t=ws.cell(row=1,column=1,value=text); t.font=Fnt(17,True,WHITE); t.alignment=Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[1].height=30
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=c2); band(ws,2,1,c2,"2E4372")
    s=ws.cell(row=2,column=1,value=sub); s.font=Fnt(10,False,"D6DEEC"); s.alignment=Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[2].height=20
def doughnut(ws,anchor,title,cats,data,colors,w=9,h=8):
    ch=DoughnutChart(); ch.title=title; ch.holeSize=52; ch.height=h; ch.width=w
    ch.add_data(data,titles_from_data=False); ch.set_categories(cats)
    s=ch.series[0]
    for i,col in enumerate(colors):
        dp=DataPoint(idx=i); dp.graphicalProperties.solidFill=col; dp.graphicalProperties.line.solidFill=WHITE; s.data_points.append(dp)
    ch.dataLabels=DataLabelList(); ch.dataLabels.showPercent=True; ch.dataLabels.numFmt='0.0%'
    ws.add_chart(ch,anchor)
def barchart(ws,anchor,title,cats,data,colors,w=10,h=8,horizontal=False,labels=True,titles_from_data=False):
    ch=BarChart(); ch.type="bar" if horizontal else "col"; ch.title=title; ch.height=h; ch.width=w
    ch.add_data(data,titles_from_data=titles_from_data); ch.set_categories(cats); ch.legend=None
    style_series(ch,colors)
    if labels:
        ch.dataLabels=DataLabelList(); ch.dataLabels.showVal=True
    ws.add_chart(ch,anchor); return ch
def linechart(ws,anchor,title,cats,data,color,w=11,h=7):
    ch=LineChart(); ch.title=title; ch.height=h; ch.width=w
    ch.add_data(data,titles_from_data=False); ch.set_categories(cats); ch.legend=None
    s=ch.series[0]; s.graphicalProperties.line.solidFill=color; s.graphicalProperties.line.width=28000
    s.marker=Marker(symbol="circle",size=6); s.smooth=False
    ch.dataLabels=DataLabelList(); ch.dataLabels.showVal=True; ch.dataLabels.numFmt='0.0%'
    ch.y_axis.numFmt='0%'; ch.y_axis.majorGridlines=None
    ws.add_chart(ch,anchor)

# ================================================================ S3 · MARKET STRUCTURE
ws=s3; widths(ws,{"A":30,"B":13,"C":12,"D":2,"E":10,"F":11,"G":11,"H":11,"I":11,"J":11,"K":11})
titleband(ws,"Slide 3 · Market structure","The headline split — how much of our volume runs through the ATC Network")
TOT=f"COUNT({C_ATC})"
r=4
kpi_card(ws,r,1,f"={TOT}","Patients (total)",NAVY)
kpi_card(ws,r,4,f"=COUNTIF({C_ATC},1)/{TOT}","Treated in ATC Network",GREEN,PCT,"share of all patients")
kpi_card(ws,r,7,f"=COUNTIF({C_ATC},0)","Patients outside ATC",SLATE_D)
kpi_card(ws,r,10,f"=COUNTIF({C_ATC},0)/{TOT}","Outside the ATC Network",AMBER,PCT,"the opportunity")
r=9
r=section(ws,r,"Site of care mix")
hdr=r; th(ws,r,["Site of care","Patients","% of total"]); r+=1
b_first=r
bmap={"ATC":GREEN,"Non-ATC: Hospital":SLATE_D,"Non-ATC: Community network":TEAL,"Non-ATC: Other":AMBER}
for b in BUCKETS:
    td(ws,r,1,b,al=LEF,fill=GREENBG if b=="ATC" else None,b=(b=="ATC"))
    td(ws,r,2,f'=COUNTIF({C_BUCKET},"{b}")',CNT)
    td(ws,r,3,f"=B{r}/$B${b_first+4}",PCT)
    r+=1
td(ws,r,1,"Total",b=True,al=LEF,fill=NAVYBG); td(ws,r,2,f"=SUM(B{b_first}:B{r-1})",CNT,b=True,fill=NAVYBG); td(ws,r,3,f"=B{r}/B{r}",PCT,b=True,fill=NAVYBG)
b_total=r
doughnut(ws,"E9","Patients by site of care",Reference(ws,min_col=1,min_row=b_first,max_row=b_first+3),
         Reference(ws,min_col=2,min_row=b_first,max_row=b_first+3),[GREEN,SLATE_D,TEAL,AMBER])
r=b_total+3
r=section(ws,r,"ATC vs non-ATC (headline split)")
th(ws,r,["Site group","Patients","% of total"]); r+=1
sp=r
for g,flag,col in [("ATC",1,GREEN),("Non-ATC",0,SLATE_D)]:
    td(ws,r,1,g,al=LEF,b=True,color=col); td(ws,r,2,f"=COUNTIF({C_ATC},{flag})",CNT); td(ws,r,3,f"=B{r}/{TOT}",PCT); r+=1
r+=1
r=section(ws,r,"ATC share by treatment-start year")
th(ws,r,["Year","Patients starting","Started at ATC","ATC share"]); r+=1
y_first=r
for y in [2021,2022,2023,2024,2025]:
    td(ws,r,1,y,INT,al=CEN); td(ws,r,2,f"=COUNTIF({C_YEAR},{y})",CNT)
    td(ws,r,3,f"=COUNTIFS({C_YEAR},{y},{C_STARTED},1)",CNT); td(ws,r,4,f"=C{r}/B{r}",PCT); r+=1
linechart(ws,"F"+str(y_first-2),"ATC share of new patients, by year",
          Reference(ws,min_col=1,min_row=y_first,max_row=y_first+4),
          Reference(ws,min_col=4,min_row=y_first,max_row=y_first+4),GREEN)
print("S3 done")

# ================================================================ S4 · PATIENT JOURNEY
ws=s4; widths(ws,{"A":20,"B":18,"C":13,"D":12,"E":2,"F":11,"G":11,"H":11,"I":11,"J":11})
titleband(ws,"Slide 4 · Patient journey","Where patients start versus where they end up — short version, where they start is where they stay")
r=4
kpi_card(ws,r,1,f'=COUNTIFS({C_FIRST},"ATC",{C_LAST},"ATC")',"Started & stayed ATC",GREEN)
kpi_card(ws,r,4,f'=COUNTIFS({C_FIRST},"Non-ATC",{C_LAST},"Non-ATC")',"Started & stayed non-ATC",SLATE_D)
kpi_card(ws,r,7,f'=COUNTIFS({C_FIRST},"Non-ATC",{C_LAST},"ATC")',"Moved INTO ATC",TEAL,CNT,"the only real switch")
kpi_card(ws,r,10,f'=COUNTIFS({C_FIRST},"ATC",{C_LAST},"Non-ATC")',"Left the ATC Network",AMBER)
r=9
r=section(ws,r,"First treatment site → last treatment site")
th(ws,r,["First site","Last site","Patients","% of total"]); r+=1
j_first=r
combos=[("Non-ATC","Non-ATC"),("ATC","ATC"),("Non-ATC","ATC"),("ATC","Non-ATC")]
labels=[]
for a,b in combos:
    lab=f"{'Started ATC' if a=='ATC' else 'Started non-ATC'} → {'ended ATC' if b=='ATC' else 'ended non-ATC'}"
    labels.append(lab)
    td(ws,r,1,("Started at an ATC" if a=="ATC" else "Started non-ATC"),al=LEF)
    td(ws,r,2,("Ended at an ATC" if b=="ATC" else "Ended non-ATC"),al=LEF)
    td(ws,r,3,f'=COUNTIFS({C_FIRST},"{a}",{C_LAST},"{b}")',CNT)
    td(ws,r,4,f"=C{r}/{TOT}",PCT); r+=1
td(ws,r,1,"Total",b=True,al=LEF,fill=NAVYBG); td(ws,r,2,None,fill=NAVYBG); td(ws,r,3,f"=SUM(C{j_first}:C{r-1})",CNT,b=True,fill=NAVYBG); td(ws,r,4,f"=C{r}/{TOT}",PCT,b=True,fill=NAVYBG)
barchart(ws,"F9","Patients by journey path",Reference(ws,min_col=1,min_row=j_first,max_row=j_first+3),
         Reference(ws,min_col=3,min_row=j_first,max_row=j_first+3),[SLATE_D,GREEN,TEAL,AMBER],horizontal=True,w=11,h=7)
r+=2
r=section(ws,r,"Treatment intensity — claims per patient, by starting site")
th(ws,r,["Started at","Patients","Avg claims / patient"]); r+=1
ci=r
for site,col in [("ATC",GREEN),("Non-ATC",SLATE_D)]:
    td(ws,r,1,("Started at ATC" if site=="ATC" else "Started at non-ATC"),al=LEF,color=col,b=True)
    td(ws,r,2,f'=COUNTIF({C_FIRST},"{site}")',CNT)
    td(ws,r,3,f'=AVERAGEIFS({C_CLAIMS},{C_FIRST},"{site}")',AV); r+=1
print("S4 done")

# ================================================================ S5 · REGIONAL PENETRATION
ws=s5; widths(ws,{"A":16,"B":18,"C":13,"D":13,"E":16,"F":2,"G":11,"H":11,"I":11,"J":11,"K":11})
titleband(ws,"Slide 5 · Regional view","Coverage by region — where the network is strong, and where it is weak")
r=4
r=section(ws,r,"Region view")
th(ws,r,["Region","Treated in ATC","Untapped (non-ATC)","Total","ATC penetration"]); r+=1
reg_first=r
for reg in REGIONS:
    td(ws,r,1,reg,al=LEF,b=True)
    td(ws,r,2,f'=COUNTIFS({C_REGION},"{reg}",{C_ATC},1)',CNT)
    td(ws,r,3,f'=COUNTIFS({C_REGION},"{reg}",{C_ATC},0)',CNT)
    td(ws,r,4,f"=B{r}+C{r}",CNT)
    td(ws,r,5,f"=B{r}/D{r}",PCT); r+=1
td(ws,r,1,"Total (mapped)",b=True,al=LEF,fill=NAVYBG)
td(ws,r,2,f"=SUM(B{reg_first}:B{r-1})",CNT,b=True,fill=NAVYBG); td(ws,r,3,f"=SUM(C{reg_first}:C{r-1})",CNT,b=True,fill=NAVYBG)
td(ws,r,4,f"=SUM(D{reg_first}:D{r-1})",CNT,b=True,fill=NAVYBG); td(ws,r,5,f"=B{r}/D{r}",PCT,b=True,fill=NAVYBG)
reg_last=r-1
# conditional formatting: data bar on untapped, color scale on penetration
ws.conditional_formatting.add(f"C{reg_first}:C{reg_last}",DataBarRule(start_type="min",end_type="max",color=AMBER))
ws.conditional_formatting.add(f"E{reg_first}:E{reg_last}",ColorScaleRule(start_type="min",start_color="F8696B",mid_type="percentile",mid_value=50,mid_color="FFEB84",end_type="max",end_color="63BE7B"))
# clustered bar treated vs untapped
ch=BarChart(); ch.type="col"; ch.grouping="clustered"; ch.title="Treated vs untapped by region"; ch.height=8; ch.width=13
ch.add_data(Reference(ws,min_col=2,max_col=3,min_row=reg_first-1,max_row=reg_last),titles_from_data=True)
ch.set_categories(Reference(ws,min_col=1,min_row=reg_first,max_row=reg_last))
ch.series[0].graphicalProperties.solidFill=GREEN; ch.series[1].graphicalProperties.solidFill=AMBER
ws.add_chart(ch,"G4")
print("S5 done")

# ================================================================ S6 · STATE SCATTER
ws=s6; widths(ws,{"A":10,"B":16,"C":16,"D":12,"E":16,"F":2})
titleband(ws,"Slide 6 · State view","Volume against coverage, by state — the targets sit top-left: lots of patients, low coverage")
r=4
r=section(ws,r,"States with 100 or more patients (the rest sit in the long tail)")
th(ws,r,["State","Treated in ATC","Untapped","Total","ATC penetration"]); r+=1
st_first=r
all_states=sorted(_m.keys())
# only include states that have patients in sample & >=100 to keep chart readable
state_tot=Counter(x["state"] for x in rows)
plot_states=[s for s in all_states if state_tot[s]>=100]
for st in plot_states:
    td(ws,r,1,st,al=CEN,b=True)
    td(ws,r,2,f'=COUNTIFS({C_STATE},"{st}",{C_ATC},1)',CNT)
    td(ws,r,3,f'=COUNTIFS({C_STATE},"{st}",{C_ATC},0)',CNT)
    td(ws,r,4,f"=B{r}+C{r}",CNT)
    td(ws,r,5,f"=IF(D{r}=0,0,B{r}/D{r})",PCT); r+=1
st_last=r-1
sc=ScatterChart(); sc.title="Untapped vs ATC penetration, by state"; sc.height=11; sc.width=16
sc.x_axis.title="ATC penetration"; sc.y_axis.title="Untapped patients"; sc.x_axis.numFmt='0%'
sc.legend=None
xref=Reference(ws,min_col=5,min_row=st_first,max_row=st_last)
yref=Reference(ws,min_col=3,min_row=st_first,max_row=st_last)
ser=Series(yref,xref,title="States"); ser.marker=Marker(symbol="circle",size=7)
ser.graphicalProperties.line.noFill=True; ser.marker.graphicalProperties.solidFill=TEAL
sc.series.append(ser)
ws.add_chart(sc,"G4")
print("S6 done")

# ================================================================ S7 · NON-ATC BY REGION
ws=s7; widths(ws,{"A":16,"B":34,"C":12,"D":13})
titleband(ws,"Slide 7 · Non-ATC accounts by region","The places our patients go instead — the biggest ones in each region",c2=8)
note(ws,3,"Genuine targets only — community networks and big multi-site systems left out.",c2=8,color=MUTE)
r=4
for reg in REGIONS:
    top=non_by_region[reg].most_common(5)
    r=section(ws,r,f"{reg}",c2=4,color=AMBER)
    th(ws,r,["Rank","Non-ATC account (parent)","Patients","% of region untapped"],fill=AMBER); r+=1
    reg_untap=f'COUNTIFS({C_REGION},"{reg}",{C_ATC},0)'
    for i,(acct_name,_) in enumerate(top,1):
        safe=acct_name.replace('"','""')
        td(ws,r,1,i,INT,al=CEN); td(ws,r,2,acct_name,al=LEF)
        td(ws,r,3,f'=COUNTIFS({C_PARENT},"{safe}",{C_REGION},"{reg}",{C_ATC},0)',CNT)
        td(ws,r,4,f'=C{r}/{reg_untap}',PCT); r+=1
    r+=1
print("S7 done")

# ================================================================ S8 · NON-ATC BY STATE
ws=s8; widths(ws,{"A":10,"B":34,"C":12,"D":13})
titleband(ws,"Slide 8 · Non-ATC accounts by state","The same account view, broken down by state — specific enough to act on",c2=8)
note(ws,3,"The eight states with the most patients outside the network.",c2=8,color=MUTE)
r=4
for st in top_states:
    top=non_by_state[st].most_common(5)
    r=section(ws,r,f"{st}  ·  region: {_m.get(st,'—')}",c2=4,color=AMBER)
    th(ws,r,["Rank","Non-ATC account (parent)","Patients","% of state untapped"],fill=AMBER); r+=1
    st_untap=f'COUNTIFS({C_STATE},"{st}",{C_ATC},0)'
    for i,(acct_name,_) in enumerate(top,1):
        safe=acct_name.replace('"','""')
        td(ws,r,1,i,INT,al=CEN); td(ws,r,2,acct_name,al=LEF)
        td(ws,r,3,f'=COUNTIFS({C_PARENT},"{safe}",{C_STATE},"{st}",{C_ATC},0)',CNT)
        td(ws,r,4,f'=C{r}/{st_untap}',PCT); r+=1
    r+=1
print("S8 done")

# ================================================================ S9 · APPENDIX
ws=s9; widths(ws,{"A":30,"B":13,"C":12,"D":2,"E":13,"F":11,"G":11,"H":11})
titleband(ws,"Slide 9 · Appendix","The extra cuts — satellite split, time to treatment, drug mix, and how the count is built")
r=4
r=section(ws,r,"How each ATC patient was matched — by a confirmed NPI, or through a satellite")
th(ws,r,["Match basis","Patients","% of ATC"]); r+=1
sat=r
ATCN=f'COUNTIF({C_ATC},1)'
for mb,col in [("NPI-confirmed",GREEN),("Roster-confirmed",TEAL),("Name-matched",AMBER)]:
    td(ws,r,1,mb,al=LEF,color=col,b=True); td(ws,r,2,f'=COUNTIF({C_MATCH},"{mb}")',CNT); td(ws,r,3,f"=B{r}/{ATCN}",PCT); r+=1
td(ws,r,1,"Total ATC",b=True,al=LEF,fill=GREENBG); td(ws,r,2,f"=SUM(B{sat}:B{r-1})",CNT,b=True,fill=GREENBG); td(ws,r,3,f"=B{r}/{ATCN}",PCT,b=True,fill=GREENBG)
doughnut(ws,"E4","ATC patients by match basis",Reference(ws,min_col=1,min_row=sat,max_row=sat+2),
         Reference(ws,min_col=2,min_row=sat,max_row=sat+2),[GREEN,TEAL,AMBER],w=8,h=7)
r+=2
r=section(ws,r,"Time from diagnosis to first treatment — about the same either way")
th(ws,r,["First site","Patients","Avg days to treatment"]); r+=1
tm=r
for site,col in [("Non-ATC",SLATE_D),("ATC",GREEN)]:
    td(ws,r,1,site,al=LEF,color=col,b=True); td(ws,r,2,f'=COUNTIF({C_FIRST},"{site}")',CNT)
    td(ws,r,3,f'=AVERAGEIFS({C_DAYS},{C_FIRST},"{site}")',INT); r+=1
r+=2
r=section(ws,r,"Drug mix — Yervoy vs Opdualag")
th(ws,r,["Drug","Patients","ATC patients","% ATC"]); r+=1
dm=r
for drug,colidx,col in [("Yervoy",C_YERVOY,GREEN),("Opdualag",C_OPD,TEAL)]:
    td(ws,r,1,drug,al=LEF,color=col,b=True); td(ws,r,2,f'=COUNTIF({colidx},1)',CNT)
    td(ws,r,3,f'=COUNTIFS({colidx},1,{C_ATC},1)',CNT); td(ws,r,4,f"=C{r}/B{r}",PCT); r+=1
barchart(ws,"E"+str(dm-1),"Patients by drug (ATC shaded)",Reference(ws,min_col=1,min_row=dm,max_row=dm+1),
         Reference(ws,min_col=2,min_row=dm,max_row=dm+1),[GREEN,TEAL],w=8,h=6)
r+=2
r=section(ws,r,"How solid the count is — every patient by how we matched them")
th(ws,r,["Match type","Patients","% of all"]); r+=1
cc=r
for cls in CLASSES:
    td(ws,r,1,cls,al=LEF); td(ws,r,2,f'=COUNTIF({C_CLASS},"{cls}")',CNT); td(ws,r,3,f"=B{r}/{TOT}",PCT); r+=1
td(ws,r,1,"Total",b=True,al=LEF,fill=NAVYBG); td(ws,r,2,f"=SUM(B{cc}:B{r-1})",CNT,b=True,fill=NAVYBG); td(ws,r,3,f"=B{r}/{TOT}",PCT,b=True,fill=NAVYBG)
r+=2
r=section(ws,r,"Community networks' share of the outside volume")
th(ws,r,["Network","Patients","% of outside"]); r+=1
cn=r; NONATCN=f'COUNTIF({C_ATC},0)'
for net in ["Independent / Other","THE US ONCOLOGY NETWORK","ONE ONCOLOGY","AMERICAN ONCOLOGY NETWORK"]:
    td(ws,r,1,net,al=LEF); td(ws,r,2,f'=COUNTIF({C_COMM},"{net}")',CNT); td(ws,r,3,f"=B{r}/{NONATCN}",PCT); r+=1
print("S9 done")

# ================================================================ PAGE 1 — ATC PATIENT COUNTS (the CEO data dump)
ws=dump
widths(ws,{"A":42,"B":12,"C":11,"D":8,"E":13,"F":2,"G":9,"H":9,"I":9,"J":9,"K":9,"L":9,"M":3})
titleband(ws,"ATC Patient Counts","One row for each ATC and how many of our patients it treats — metastatic melanoma, Yervoy and Opdualag, 2021 to 2025",c2=13)
note(ws,3,"A row per ATC with its patient count. The analysis is in the tabs after this one.",c2=13,color=MUTE)
NPAR=len(atc_parents); a_first=15; a_last=a_first+NPAR-1
kpi_card(ws,5,2,f"={TOT}","Patients in all",NAVY,CNT,"metastatic melanoma")
kpi_card(ws,5,5,f"={ATCN}","Treated in the network",GREEN,CNT,"about 46 percent")
kpi_card(ws,5,8,f"=COUNTA($A${a_first}:$A${a_last})","Authorized centers",TEAL,INT,"one row each, below")
kpi_card(ws,5,11,f"={NONATCN}","Treated outside it",AMBER,CNT,"where the room to grow is")
# one-line read (live)
ws.merge_cells("A10:M10"); band(ws,10,1,13,NAVYBG)
oc=ws.cell(10,1,f'="A little over half of our patients — "&TEXT({NONATCN},"#,##0")&" of "&TEXT({TOT},"#,##0")&" ("&TEXT({NONATCN}/{TOT},"0%")&") — are treated outside the ATC Network. Tabs S5 to S8 show where they go."')
oc.font=Fnt(10.5,True,NAVY); oc.alignment=LEF; ws.row_dimensions[10].height=24
r=12
r=section(ws,r,"Our patients at each ATC, most to fewest",c2=5)
th(ws,r,["ATC account (parent)","Patients","Share of ATC","State","Treatment claims"]); r+=1
for p in atc_parents:
    safe=p.replace('"','""')
    td(ws,r,1,p,al=LEF)
    td(ws,r,2,f'=COUNTIFS({C_PARENT},"{safe}",{C_ATC},1)',CNT)
    td(ws,r,3,f"=B{r}/{ATCN}",PCT)
    td(ws,r,4,f'=IFERROR(INDEX({C_STATE},MATCH("{safe}",{C_PARENT},0)),"")',al=CEN)
    td(ws,r,5,f'=SUMIFS({C_CLAIMS},{C_PARENT},"{safe}",{C_ATC},1)',CNT); r+=1
a_last=r-1
td(ws,r,1,"All ATCs together",b=True,al=LEF,fill=GREENBG)
td(ws,r,2,f"=SUM(B{a_first}:B{a_last})",CNT,b=True,fill=GREENBG)
td(ws,r,3,f"=B{r}/{ATCN}",PCT,b=True,fill=GREENBG); td(ws,r,4,None,fill=GREENBG)
td(ws,r,5,f"=SUM(E{a_first}:E{a_last})",CNT,b=True,fill=GREENBG)
ws.conditional_formatting.add(f"B{a_first}:B{a_last}",DataBarRule(start_type="min",end_type="max",color=GREEN))
r+=2
r=section(ws,r,"About the data",c2=5)
MSUM=[
 "McKesson (Compile) claims, 2021 to 2025 — 16,246 metastatic melanoma patients on Yervoy or Opdualag.",
 "An ATC is counted at the parent, so satellites count. We match on the NPI first, then the parent name, and count each patient once at their main site.",
 "7,501 (about 46 percent) are treated in the network; the other 8,745 are not.",
 "Counts by organization only — round anything under about 11 patients before it's shared outside. Full detail on the Methodology tab.",
]
for line in MSUM:
    r=note(ws,r,"•  "+line,c2=8,color=INK,size=9); ws.row_dimensions[r-1].height=max(14,12*(len(line)//95+1))
r+=1
lnk=ws.cell(r,1,"▶  See the full analysis (dashboard)"); lnk.font=Fnt(11,True,"0563C1"); lnk.hyperlink="#'Cover'!A1"
lnk.fill=Fill(CARD); lnk.border=BORDER; lnk.alignment=LEF
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=3)
ws.freeze_panes=f"A{a_first}"
print("page-1 dump done")

# ================================================================ METHODOLOGY
ws=meth; widths(ws,{"A":28,"B":96})
titleband(ws,"Methodology & Sources","How ATC vs non-ATC is defined, the corrections applied, and data sensitivity",c2=2)
r=4; th(ws,r,["Item","Detail"]); r+=1
M=[("Business question","Metastatic melanoma patients on Yervoy or Opdualag, 2021–2025: what share are treated at an Authorized Treatment Center vs a non-ATC site, and where is the opportunity?"),
   ("Data source","McKesson (Compile) medical claims — COMPILE_CLAIMS.OPEN_CLAIMS.IOV2501_MEDICAL_CLAIMS."),
   ("Time window","Date of service 2021-01-01 to 2025-12-31."),
   ("Population","Metastatic melanoma (C43 with C77/C78/C79) treated with Yervoy (J9228 / NDCs 00003232711, 00003232822) or Opdualag (J9298 / NDC 00003712511)."),
   ("Total patients","16,246 distinct patients."),
   ("ATC definition","A site rolls up to its authorized parent. Match order: (1) NPI on the ATC roster; (2) roster-gap-corrected parents; (3) authorized parent name, if that parent spans ≤2 states. Community networks are carved out as non-ATC."),
   ("Patient assignment","Each patient counted once, at the site with the most treatment claims (primary site)."),
   ("ATC share","≈46.2% (7,501 of 16,246). ~54% treated outside the ATC Network."),
   ("Match-basis confidence","NPI-confirmed is the firm floor; name-matched rests on parent-name matching (a close estimate). See S9 satellite split and the Classification table."),
   ("Roster gap correction (2026-07-17)","Four authorized orgs were missing from the roster and scored non-ATC. Confirmed against Infinity's ATC master and moved to ATC: City of Hope, NYU Langone, Ohio State (Wexner), Hoag. Effect: +566 patients, 6,935→7,501 (42.7%→46.2%)."),
   ("Journey alignment (2026-07-21)","Journey and year-trend use the SAME ATC definition as the headline. The honest read is retention, not migration."),
   ("How the workbook works","Every slide reads from the Patient Data tab. Swap in the real export and all the numbers, KPIs and charts follow — nothing to rebuild by hand."),
   ("Handle with care","Counts by organization only — no patient-level detail on the slides. Anything under about 11 patients (mostly the long tail outside the network) should be rounded or held back before it leaves the building."),
   ("This file","Built 22 July 2026 with sample data that matches the deck, so it reads true out of the box. Put the real export in before it goes anywhere."),
]
for k,v in M:
    a=td(ws,r,1,k,al=TOP,b=True,fill=PAPER); a.alignment=TOP
    b=td(ws,r,2,v,al=TOP); b.alignment=TOP
    ws.row_dimensions[r].height=max(26,13*(len(v)//96+1)); r+=1
print("methodology done")

# ================================================================ COVER (built last; references computed cells)
ws=cover; widths(ws,{"A":3,"B":20,"C":13,"D":13,"E":13,"F":13,"G":13,"H":13,"I":13,"J":13,"K":13,"L":13,"M":3})
band(ws,1,1,13,NAVY); ws.merge_cells("A1:M1"); ws.row_dimensions[1].height=8
ws.merge_cells("B3:L3"); t=ws.cell(3,2,"ATC SITE OF CARE"); t.font=Fnt(30,True,NAVY); t.alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[3].height=40
ws.merge_cells("B4:L4"); st_=ws.cell(4,2,"Where our metastatic melanoma patients actually get treated — Yervoy and Opdualag, 2021 to 2025"); st_.font=Fnt(12,False,MUTE);
ws.merge_cells("B5:L5"); ex=ws.cell(5,2,"Every number reads from the Patient Data tab."); ex.font=Fnt(10,False,TEAL,True)
# KPI ROW
r=8
kpi_card(ws,r,2,f"={TOT}","Patients in all",NAVY)
kpi_card(ws,r,5,f"=COUNTIF({C_ATC},1)/{TOT}","In the ATC Network",GREEN,PCT)
kpi_card(ws,r,8,f"=COUNTIF({C_ATC},0)","Treated outside it",AMBER)
kpi_card(ws,r,11,f"=COUNTA('ATC Patient Counts'!$A${a_first}:$A${a_last})","Authorized centers",TEAL,INT)
# site-of-care chart (reads the S3 bucket table)
doughnut(ws,"B13","Patients by site of care",
         Reference(s3,min_col=1,min_row=12,max_row=15),Reference(s3,min_col=2,min_row=12,max_row=15),[GREEN,SLATE_D,TEAL,AMBER],w=9,h=8)
# interactive picker
ws.merge_cells("H13:L13"); h=ws.cell(13,8,"Pick a region to focus on:"); h.font=Fnt(11,True,NAVY)
dv=DataValidation(type="list",formula1='"All,West,Central,Great Lakes,Ohio Valley,Southeast,Northeast"',allow_blank=False)
ws.add_data_validation(dv); ws["L14"]="All"; ws["L14"].fill=Fill(INPUT); ws["L14"].font=Fnt(11,True,NAVY); ws["L14"].alignment=CEN; ws["L14"].border=BORDER
dv.add(ws["L14"])
ws["H14"]="Region:"; ws["H14"].font=Fnt(10,True,MUTE); ws["H14"].alignment=RGT
ws.merge_cells("H15:L15"); ws.cell(15,8,"The region you picked").font=Fnt(10,True,TEAL)
kpi_card(ws,16,8,f'=IF($L$14="All",{TOT},COUNTIF({C_REGION},$L$14))',"Patients there",NAVY,CNT,w=2)
kpi_card(ws,16,10,f'=IF($L$14="All",COUNTIF({C_ATC},0),COUNTIFS({C_REGION},$L$14,{C_ATC},0))',"Treated outside the network there",AMBER,CNT,w=3)
# nav
r=23
ws.cell(r,2,"Jump to a slide").font=Fnt(12,True,NAVY); r+=1
nav=[("ATC Patient Counts (page 1)",dump.title),("S3 · Market Structure",s3.title),("S4 · Patient Journey",s4.title),
     ("S5 · Regional Penetration",s5.title),("S6 · State Scatter",s6.title),("S7 · Non-ATC by Region",s7.title),
     ("S8 · Non-ATC by State",s8.title),("S9 · Appendix",s9.title),("Methodology",meth.title),
     ("Patient Data",pdata.title),("Load Data (SQL)",loadsql.title)]
cc=2
for lab,target in nav:
    cell=ws.cell(r,cc,"▶ "+lab); cell.font=Fnt(10,True,"0563C1"); cell.hyperlink=f"#'{target}'!A1"
    cell.fill=Fill(CARD); cell.border=BORDER; cell.alignment=LEF
    ws.merge_cells(start_row=r,start_column=cc,end_row=r,end_column=cc+2)
    cc+=3
    if cc>11: cc=2; r+=1
r+=2
note(ws,r,"Sample data — swap in the real export on the Patient Data tab and everything follows.",c2=12,color=MUTE)

# paper background on the Cover only (keeps slide tables clean white)
paper(cover,rows_=40,cols_=14)

# freeze + calc
for name in wb.sheetnames:
    w_=wb[name]
    if name not in ("Cover","ATC Patient Counts","Patient Data","Load Data (SQL)","Region Map","Methodology"):
        w_.freeze_panes="A3"
# order: data dump first (what the CEO asked for), then the analysis, then the plumbing
order=["ATC Patient Counts","Cover","S3 · Market Structure","S4 · Patient Journey","S5 · Regional Penetration",
       "S6 · State Scatter","S7 · Non-ATC by Region","S8 · Non-ATC by State","S9 · Appendix",
       "Methodology","Patient Data","Load Data (SQL)","Region Map"]
wb._sheets=[wb[n] for n in order]
wb.active=0
from openpyxl.workbook.properties import CalcProperties
wb.calculation=CalcProperties(calcId=0,fullCalcOnLoad=True)
wb.save(OUT)
print("SAVED:",OUT,"| sheets:",len(wb.sheetnames))
